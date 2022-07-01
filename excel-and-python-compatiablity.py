from openpyxl import Workbook, load_workbook 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
    "Alex": {
        "Bingus": 100,
        "Bongus": 99,
        "metagaming": 0,
        "sponsorship": 50
    },
    "Tanya": {
        "Bingus": 99,
        "Bongus": 89,
        "metagaming": 30,
        "sponsorship": 200
    },
    "Hot Floppa": {
        "Bingus": 100,
        "Bongus": 99,
        "metagaming": 0,
        "sponsorship": 50
    },
    'Sogga': {
        "Bingus": 100,
        "Bongus": 99,
        "metagaming": 0,
        "sponsorship": 50
    },
}


wb = Workbook()
ws = wb.active
ws.title = "People and Pokemon"

headings = ['Name'] + list(data['Alex'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person]+ grades)
wb.save("People and Pokemon bingus.xlsx")

# ws = wb.active

# for row in range(1,11):
#     for col in range (1,5):
#         char = get_column_letter(col)
#         ws[char + str(row)] = char + str(row)

# ws.unmerge_cells("A1:D1")

# ws.move_range('C1:d5', rows=8, cols=5)
# ws.insert_row(row)
# ws.instert_cols(colounm) 





